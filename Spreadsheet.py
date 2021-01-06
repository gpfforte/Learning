from tkinter import *
from openpyxl.workbook import Workbook

from openpyxl import load_workbook

filename= "countries_red.xlsx"
root = Tk()
root.title('Learning Python')
root.geometry("1000x700")
root.iconphoto(True, PhotoImage(file='Images/Icona.png'))

#wb=Workbook()
wb=load_workbook(filename)
ws=wb.active
column_a=ws["A"]
column_b=ws["B"]



def get_a():
    lista=""
    altra_lista=[]
    my_listbox.delete(0,END)
    label_a.config(text=lista)
    for cella in column_a:
        lista=f"{lista+str(cella.value)}\n"
        altra_lista.append(str(cella.value))
        print(cella.value)
    label_a.config(text=lista)
    for riga in altra_lista:
        my_listbox.insert(END, riga)

btn_a=Button(root, text="Column A", command=get_a)
btn_a.pack(pady=10, padx=10)

label_a=Label(root, text="")
label_a.pack(pady=10)

def get_b():
    for cella in column_b:
        print(cella.value)

def select(event):
    my_label.config(text = my_listbox.get(ANCHOR))
btn_b=Button(root, text="Column B", command=get_b)
btn_b.pack(pady=10, padx=10)

my_listbox=Listbox(root, width=50)
my_listbox.pack(pady=20)
my_listbox.bind("<ButtonRelease-1>", select)

my_label=Label(root, text="Select Item...", font=("Times New Roman", 20))
my_label.pack(pady=20)



# # Add valuee
# ws["A13"]="ITA"
# ws["B13"]="Italia"
#
# # Save file
#
# wb.save(filename)

#print(column_a)

root.mainloop()
