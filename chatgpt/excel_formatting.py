import os

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Setta la working directory al path dello script
abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)
# Carica il file Excel
file_path = "tuo_file_excel.xlsx"
workbook = openpyxl.load_workbook(file_path)

# Seleziona il foglio di lavoro desiderato
sheet = workbook.active

# Specifica le colonne su cui applicare la formattazione condizionale
colonne_da_formattare = [1, 2, 3]  # Aggiungi le colonne desiderate
colonne_da_formattare = ["A", "B", "C"]  # Aggiungi le colonne desiderate

# Configura la formattazione condizionale
fill = PatternFill(start_color="FFFFFF", end_color="FFFF00", fill_type="solid")
cell_rule = CellIsRule(operator="between", formula=[10, 20], fill=fill)

for colonna in colonne_da_formattare:
    # for row in sheet.iter_rows(min_col=1, min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
    for row in sheet.iter_rows(min_col=1, min_row=1):
        for cell in row:
            print(colonna)
            print(cell.column_letter)
            print(cell.column_letter == colonna)
            # if cell.column == colonna:
            if cell.column_letter == colonna:
                sheet.conditional_formatting.add(cell.coordinate, cell_rule)

# Salva le modifiche nel file Excel
workbook.save("file_excel_con_formattazione.xlsx")

# Chiudi il file Excel
workbook.close()
