from tkinter import *
from openpyxl.workbook import Workbook

from openpyxl import load_workbook
# from PDFWriter import PDFWriter

filename = "Cities.xlsx"
root = Tk()
root.title('Learning Python')
root.geometry("1000x700")
# root.iconphoto(True, PhotoImage(file='Images/Icona.png'))

# wb=Workbook()
wb = load_workbook(filename)
ws = wb.active
column_a = ws["A"]
column_b = ws["B"]

print (wb.sheetnames)



def find_specific_cell():
    for row in range(1, ws.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if ws[cell_name].value == "Somma":
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, ws[cell_name].value))
                return cell_name


def inserisci_somma():
    print(len(column_b))
    cella_trovata= find_specific_cell()
    print (cella_trovata)
    riga_somma=int(cella_trovata[1:])
    cella_con_somma="B"+str(riga_somma)
    ws[cella_con_somma] = "=SUM(B2:B"+str(riga_somma-1)+")"

    # ws.cell(len(column_b),2,"=SUM(B2:B4)")


def salva():
    wb.save(filename)


def get_a():
    lista = ""
    altra_lista = []
    my_listbox.delete(0, END)
    label_a.config(text=lista)
    for cella in column_a:
        lista = f"{lista + str(cella.value)}\n"
        altra_lista.append(str(cella.value))
        print(cella.value)
    label_a.config(text=lista)
    for riga in altra_lista:
        my_listbox.insert(END, riga)


btn_a = Button(root, text="Column A", command=get_a)
btn_a.pack(pady=10, padx=10)

btn_prt = Button(root, text="Inserisci Somma", command=inserisci_somma)
btn_prt.pack(pady=10, padx=10)

btn_save = Button(root, text="Salva", command=salva)
btn_save.pack(pady=10, padx=10)

label_a = Label(root, text="")
label_a.pack(pady=10)


def get_b():
    for cella in column_b:
        print(cella.value)


def select(event):
    my_label.config(text=my_listbox.get(ANCHOR))


btn_b = Button(root, text="Column B", command=get_b)
btn_b.pack(pady=10, padx=10)

my_listbox = Listbox(root, width=50)
my_listbox.pack(pady=20)
my_listbox.bind("<ButtonRelease-1>", select)

my_label = Label(root, text="Select Item...", font=("Times New Roman", 20))
my_label.pack(pady=20)

# # Add valuee
# ws["A13"]="ITA"
# ws["B13"]="Italia"
#
# # Save file
#
# wb.save(filename)

# print(column_a)

root.mainloop()

